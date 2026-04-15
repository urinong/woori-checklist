"""체크리스트 엔진 — 물품사양서로부터 자가진단 체크리스트 생성

핵심: 판정(승인/반려)은 내리지 않는다. 체크 항목 + 설명 + 액션만 제공.
"""
from dataclasses import dataclass, field, asdict
from typing import List, Dict, Any, Optional
from openpyxl import load_workbook
import re

from .pdf_parser import ProductSpec, Ingredient
from config import DATA_FILES


# ---------------------------------------------------------------------------
# 데이터 클래스
# ---------------------------------------------------------------------------

@dataclass
class ChecklistItem:
    category: str            # 필요서류 / 기재점검 / 원재료주의 / 이슈대비
    status: str              # required / submitted / missing / warning / ok
    title: str
    description: str         # 신입직원용 설명
    action: str              # 실무자 행동
    regulation_ref: str = ""
    icon: str = ""

    def as_dict(self) -> Dict[str, Any]:
        return asdict(self)


@dataclass
class ChecklistResult:
    product_info: Dict[str, Any] = field(default_factory=dict)
    required_documents: List[ChecklistItem] = field(default_factory=list)
    spec_check: List[ChecklistItem] = field(default_factory=list)
    ingredient_warnings: List[ChecklistItem] = field(default_factory=list)
    potential_issues: List[ChecklistItem] = field(default_factory=list)
    request_to_producer: List[str] = field(default_factory=list)

    def as_dict(self) -> Dict[str, Any]:
        return {
            "product_info": self.product_info,
            "required_documents": [i.as_dict() for i in self.required_documents],
            "spec_check": [i.as_dict() for i in self.spec_check],
            "ingredient_warnings": [i.as_dict() for i in self.ingredient_warnings],
            "potential_issues": [i.as_dict() for i in self.potential_issues],
            "request_to_producer": self.request_to_producer,
        }


# ---------------------------------------------------------------------------
# 규정 데이터 로더 (캐시)
# ---------------------------------------------------------------------------

_cache: Dict[str, Any] = {}


def _load_excel(key: str) -> Any:
    if key in _cache:
        return _cache[key]
    wb = load_workbook(DATA_FILES[key], data_only=True)
    _cache[key] = wb
    return wb


def _rows(wb, sheet) -> List[tuple]:
    ws = wb[sheet]
    it = ws.iter_rows(values_only=True)
    header = next(it, None)
    return list(it)


def _dict_rows(wb, sheet) -> List[Dict[str, Any]]:
    ws = wb[sheet]
    it = ws.iter_rows(values_only=True)
    header = list(next(it, ()))
    out = []
    for r in it:
        out.append({header[i]: (r[i] if i < len(r) else None) for i in range(len(header))})
    return out


# ---------------------------------------------------------------------------
# 진단 엔진
# ---------------------------------------------------------------------------

class ChecklistEngine:

    def __init__(self):
        self.docs = _dict_rows(_load_excel("제출서류"), "제출서류")
        self.display_fields = _dict_rows(_load_excel("표시사항"), "필수표시항목")
        gmo_wb = _load_excel("GMO")
        self.gmo_crops = _dict_rows(gmo_wb, "GMO_농산물")
        self.gmo_check = _dict_rows(gmo_wb, "GMO_확인필요원료")
        imp_wb = _load_excel("수입허용")
        self.imp_spices = [r["원료명"] for r in _dict_rows(imp_wb, "향신료") if r.get("원료명")]
        self.imp_herbs = [r["원료명"] for r in _dict_rows(imp_wb, "약재") if r.get("원료명")]
        self.imp_oils = [r["원료명"] for r in _dict_rows(imp_wb, "기름류") if r.get("원료명")]
        self.bunwa = _dict_rows(_load_excel("분과"), "식품유형_분과매핑")

    # ---- public ---------------------------------------------------------
    def diagnose(self, spec: ProductSpec) -> ChecklistResult:
        result = ChecklistResult()
        result.product_info = self._product_info(spec)

        # 조건 플래그 수집
        flags = self._collect_flags(spec)
        result.product_info["분과"] = flags.get("분과", "")

        # 1) 필요 서류 체크리스트
        result.required_documents = self._required_documents(spec, flags)

        # 2) 물품사양서 기재 점검
        result.spec_check = self._spec_check(spec)

        # 3) 원재료 주의사항
        result.ingredient_warnings = self._ingredient_warnings(spec, flags)

        # 4) 출하 승인 시 예상 이슈
        result.potential_issues = self._potential_issues(spec, flags)

        # 5) 생산자 요청사항 텍스트
        result.request_to_producer = self._build_request(spec, result)

        return result

    # ---- product info ---------------------------------------------------
    def _product_info(self, spec: ProductSpec) -> Dict[str, Any]:
        return {
            "제품명": spec.제품명,
            "식품유형": spec.식품유형,
            "제조원": spec.제조원,
            "판매원": spec.판매원,
            "생산자": spec.생산자,
            "교구": spec.교구,
            "회원구분": spec.회원구분,
            "내용량": spec.내용량,
            "소비기한": spec.소비기한,
            "사용용수": spec.사용용수,
            "설비위생": spec.설비위생,
            "작성일시": spec.작성일시,
            "ingredients_count": len(spec.ingredients),
        }

    # ---- flag collection ------------------------------------------------
    def _collect_flags(self, spec: ProductSpec) -> Dict[str, Any]:
        flags: Dict[str, Any] = {
            "사용용수": (spec.사용용수 or "").strip(),
            "원료유형": set(),
            "소금종류": "",
            "분과": "",
            "영양성분의무": False,
            "GMO원료": [],      # [(ingredient, 확인사항)]
            "수입원료_허용없음": [],  # ingredients
            "배합비_공란": [],
            "천일염포함": False,
            "수산물포함": False,
            "기름류포함": False,
        }

        # 식품유형 → 분과
        f = (spec.식품유형 or "").strip()
        for row in self.bunwa:
            t = (row.get("식품유형") or "").strip()
            if not t:
                continue
            t_base = t.rstrip("류")  # '과자류' → '과자'
            if t in f or (t_base and t_base in f):
                flags["분과"] = row.get("분과구분", "")
                if (row.get("영양성분표시") or "") == "의무":
                    flags["영양성분의무"] = True
                break

        # 원재료 검사
        for ing in spec.ingredients:
            name = (ing.name or "").strip()
            origin = (ing.origin or "").strip()
            ratio = (ing.ratio or "").strip()

            # 배합비 공란
            if not ratio:
                flags["배합비_공란"].append(name)

            # 수산물
            수산_키워드 = ["가자미", "고등어", "갈치", "조기", "명태", "황태", "오징어",
                          "멸치", "다시마", "미역", "김", "새우", "게", "문어", "낙지", "연육"]
            if any(k in name for k in 수산_키워드):
                flags["수산물포함"] = True
                flags["원료유형"].add("수산물")

            # 소금/천일염
            if "천일염" in name:
                flags["천일염포함"] = True
                flags["소금종류"] = "천일염"
            elif "소금" in name and not flags["소금종류"]:
                flags["소금종류"] = "미특정"

            # 기름류
            기름_키워드 = ["유", "기름"]
            if ("현미유" in name or "포도씨유" in name or "해바라기씨유" in name
                    or "올리브유" in name or "참기름" in name or "들기름" in name
                    or "팜유" in name):
                flags["기름류포함"] = True
                if "참기름" in name or "들기름" in name:
                    flags["원료유형"].add("기름류")

            # GMO 위험 키워드
            for gmo in self.gmo_check:
                kw = (gmo.get("키워드") or "").strip()
                if kw and kw in name:
                    flags["GMO원료"].append((ing, gmo.get("확인사항", ""), gmo.get("설명_긴", "")))
                    flags["원료유형"].add("GMO위험")
                    break

            # 수입원료 허용 여부
            if self._is_imported(origin):
                if not self._is_allowed_import(name):
                    flags["수입원료_허용없음"].append(ing)

        return flags

    @staticmethod
    def _is_imported(origin: str) -> bool:
        if not origin:
            return False
        o = origin.strip()
        if "국산" in o or "국내" in o:
            return False
        수입국 = ["이스라엘", "태국", "중국", "베트남", "미국", "일본", "인도", "필리핀",
                 "에콰도르", "페루", "칠레", "브라질", "프랑스", "이탈리아", "스페인",
                 "독일", "러시아", "호주", "뉴질랜드", "캐나다", "수입"]
        return any(k in o for k in 수입국)

    def _is_allowed_import(self, name: str) -> bool:
        pools = self.imp_spices + self.imp_herbs + self.imp_oils
        for p in pools:
            if p and (p in name or name in p):
                return True
        return False

    # ---- 필요서류 --------------------------------------------------------
    def _required_documents(self, spec: ProductSpec, flags: Dict[str, Any]) -> List[ChecklistItem]:
        items: List[ChecklistItem] = []
        for d in self.docs:
            필수 = (d.get("필수여부") or "").strip()
            조건필드 = (d.get("조건_필드") or "").strip() if d.get("조건_필드") else ""
            조건값 = (d.get("조건_값") or "").strip() if d.get("조건_값") else ""
            name = d.get("서류명", "")
            desc_short = d.get("설명_짧은", "") or ""
            desc_long = d.get("설명_긴", "") or ""
            req_how = d.get("요청방법", "") or ""

            include = False
            if 필수 == "필수":
                include = True
            elif 필수 == "조건부":
                include = self._condition_met(조건필드, 조건값, spec, flags)

            if include:
                items.append(ChecklistItem(
                    category="필요서류",
                    status="required",
                    title=name,
                    description=desc_long or desc_short,
                    action=req_how,
                    icon="📄",
                ))
        return items

    def _condition_met(self, 필드: str, 값: str, spec: ProductSpec, flags: Dict[str, Any]) -> bool:
        if 필드 == "사용용수":
            return 값 in (spec.사용용수 or "")
        if 필드 == "원료유형":
            return 값 in flags["원료유형"]
        if 필드 == "소금종류":
            return 값 in (flags["소금종류"] or "")
        return False

    # ---- 기재 점검 -------------------------------------------------------
    def _spec_check(self, spec: ProductSpec) -> List[ChecklistItem]:
        items: List[ChecklistItem] = []
        # 매핑: 필수표시항목명 → spec 필드
        field_map = {
            "제품명": "제품명",
            "식품의유형": "식품유형",
            "제조원": "제조원",
            "판매원": "판매원",
            "소비기한": "소비기한",
            "소비기한표시": "소비기한표시",
            "내용량": "내용량",
            "원재료및함량": "원재료표기",
            "알레르기물질": "알레르기물질",
            "용기포장재질": "용기포장재질",
            "품목보고번호": "품목보고번호",
            "소비자상담실": "소비자상담실",
            "바코드": "바코드",
            "보관방법": "보관방법",
            "반품및교환": "반품교환",
            "영양성분표시": "영양성분표시",
            "생산관리": "생산관리",
            "업체담당자": "업체담당자",
            "사용용수": "사용용수",
            "설비위생": "설비위생",
        }
        for row in self.display_fields:
            항목 = row.get("항목명", "")
            desc_short = row.get("설명_짧은", "") or ""
            desc_long = row.get("설명_긴", "") or ""
            attr = field_map.get(항목)
            if not attr:
                continue
            val = getattr(spec, attr, "") or ""
            val = str(val).strip()
            if val:
                items.append(ChecklistItem(
                    category="기재점검",
                    status="ok",
                    title=f"{항목} — 기재됨",
                    description=f"현재값: {val}",
                    action="",
                    icon="✅",
                ))
            else:
                items.append(ChecklistItem(
                    category="기재점검",
                    status="missing",
                    title=f"{항목} — 공란",
                    description=desc_long or desc_short,
                    action=f"물품사양서 '{항목}' 항목 기재를 요청하세요",
                    icon="⚠",
                ))
        return items

    # ---- 원재료 주의사항 -------------------------------------------------
    def _ingredient_warnings(self, spec: ProductSpec, flags: Dict[str, Any]) -> List[ChecklistItem]:
        items: List[ChecklistItem] = []

        # GMO 위험 원료
        for ing, check, desc_long in flags["GMO원료"]:
            items.append(ChecklistItem(
                category="원재료주의",
                status="warning",
                title=f"Non-GMO 확인 필요 — {ing.name}",
                description=desc_long or f"{ing.name}은(는) GMO 대상 원료로 확인이 필요합니다.",
                action=f"제조원에 '{ing.name} Non-GMO 확인서 또는 IP인증서'를 요청하세요",
                regulation_ref="생산규정 5조",
                icon="⚠",
            ))

        # 수입원료 허용 목록 미포함
        for ing in flags["수입원료_허용없음"]:
            items.append(ChecklistItem(
                category="원재료주의",
                status="warning",
                title=f"수입원료 허용 여부 확인 — {ing.name} ({ing.origin})",
                description=(
                    f"{ing.name}은(는) 수입 원료({ing.origin})인데, "
                    f"우리농 허용 수입원료 목록(향신료 18종, 약재 4종, Non-GMO 기름류)에 포함되지 않습니다. "
                    f"원부재료(과일가공품 등)로 분류되면 수입유기농 인증이 필요하고, "
                    f"첨가물/향신료로 분류되면 다른 기준이 적용됩니다."
                ),
                action="전국물품위원회 논의가 필요한 사항입니다. 유기인증 여부와 분류 기준을 먼저 확인하세요.",
                regulation_ref="생산규정 6조 (수입원료)",
                icon="⚠",
            ))

        # 배합비 공란
        if flags["배합비_공란"]:
            names = ", ".join(flags["배합비_공란"])
            items.append(ChecklistItem(
                category="원재료주의",
                status="warning",
                title=f"배합비(%) 미기재 — {len(flags['배합비_공란'])}건",
                description=(
                    f"{names} 의 배합비(%)가 비어있습니다. "
                    f"우리농은 전 원료의 배합비가 기재되어 있어야 합니다."
                ),
                action="제조원에 '전 원재료 배합비(%) 기재'를 요청하세요",
                regulation_ref="물품사양서 작성요령",
                icon="⚠",
            ))

        # 천일염
        if flags["천일염포함"]:
            items.append(ChecklistItem(
                category="원재료주의",
                status="warning",
                title="천일염 사용 — 방사능 시험성적서 필요",
                description="천일염을 사용하는 경우 바다 환경의 방사능 영향을 확인하기 위한 별도의 방사능 시험성적서가 필요합니다.",
                action="제조원에 '천일염 방사능 시험성적서'를 요청하세요",
                icon="⚠",
            ))
        elif flags["소금종류"] == "미특정":
            items.append(ChecklistItem(
                category="원재료주의",
                status="warning",
                title="소금 종류 확인 필요",
                description="원재료에 '소금'으로만 표기되어 있습니다. 천일염/정제염/구운소금/자염/죽염 중 어떤 것인지 확인이 필요합니다. 천일염이면 방사능 성적서도 필요합니다.",
                action="제조원에 소금 종류 구체 기재를 요청하세요",
                icon="⚠",
            ))

        return items

    # ---- 이슈 대비 -------------------------------------------------------
    def _potential_issues(self, spec: ProductSpec, flags: Dict[str, Any]) -> List[ChecklistItem]:
        items: List[ChecklistItem] = []

        # 원재료 분류 경계 사례 (수입원료 허용없음 건)
        for ing in flags["수입원료_허용없음"]:
            items.append(ChecklistItem(
                category="이슈대비",
                status="warning",
                title=f"{ing.name} — 원료 분류 경계 사례",
                description=(
                    f"{ing.name}({ing.origin})이 원부재료(과일가공품 등)인지 "
                    f"첨가물/향신료인지에 따라 적용 기준이 달라집니다. "
                    f"전국물품위원회 논의가 필요할 수 있습니다."
                ),
                action="분류 기준과 유기인증 여부를 사전 확인하세요",
                icon="💡",
            ))

        # 수산물 + 양식/자연산 미명시
        if flags["수산물포함"]:
            raw = spec.raw_text
            if "자연산" not in raw and "양식" not in raw:
                items.append(ChecklistItem(
                    category="이슈대비",
                    status="warning",
                    title="양식/자연산 여부 미기재",
                    description=(
                        "수산물 원료가 자연산인지 양식인지 물품사양서에 명시되어 있지 않습니다. "
                        "양식 수산물인 경우 항생물질 사용 여부 확인이 필요합니다."
                    ),
                    action="제조원에 '자연산/양식 여부 및 양식 시 항생제 사용 여부'를 요청하세요",
                    regulation_ref="수산분과 생산기준 제2항",
                    icon="💡",
                ))

        # 영양성분 의무 표시 대상 확인
        if flags["영양성분의무"] and not (spec.영양성분표시 or "").strip():
            items.append(ChecklistItem(
                category="이슈대비",
                status="warning",
                title="영양성분 표시 의무 대상",
                description=f"{spec.식품유형}은(는) 영양성분 표시 의무 대상입니다. 제품 포장지에 영양성분 표가 인쇄되어 있는지 확인이 필요합니다.",
                action="포장지 영양성분 표시 유무를 확인하세요",
                icon="💡",
            ))

        return items

    # ---- 생산자 요청사항 텍스트 ------------------------------------------
    def _build_request(self, spec: ProductSpec, result: ChecklistResult) -> List[str]:
        reqs: List[str] = []

        # 배합비
        if any(i.title.startswith("배합비") for i in result.ingredient_warnings):
            reqs.append("물품사양서 원재료 배합비(%) 전체 기재")

        # 기재 공란
        for it in result.spec_check:
            if it.status == "missing":
                item_name = it.title.split(" — ")[0]
                reqs.append(f"물품사양서 '{item_name}' 항목 기재")

        # 원재료 주의사항
        for it in result.ingredient_warnings:
            if it.title.startswith("Non-GMO"):
                reqs.append(it.title.replace("Non-GMO 확인 필요 — ", "") + " Non-GMO 확인서 제출")
            elif it.title.startswith("수입원료"):
                name = it.title.replace("수입원료 허용 여부 확인 — ", "").split(" (")[0]
                reqs.append(f"{name} 유기인증 여부 및 분류 근거 확인")
            elif it.title.startswith("천일염"):
                reqs.append("천일염 방사능 시험성적서 제출")
            elif it.title.startswith("소금 종류"):
                reqs.append("소금 종류(천일염/정제염 등) 구체 기재")

        # 이슈대비 중 수산물 양식/자연산
        for it in result.potential_issues:
            if "양식/자연산" in it.title:
                reqs.append("수산물 원료 자연산/양식 여부 기재 (양식 시 항생제 사용 여부 포함)")

        return reqs
