"""체크리스트 결과 → 엑셀 보고서"""
from typing import Optional
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .checklist_engine import ChecklistResult
from config import OUTPUT_DIR


HEADER_FILL = PatternFill("solid", fgColor="2E7D32")
SECTION_FILL = PatternFill("solid", fgColor="C8E6C9")
WHITE = Font(color="FFFFFF", bold=True, size=12)
BOLD = Font(bold=True, size=11)
WRAP = Alignment(wrap_text=True, vertical="top")
BORDER = Border(*[Side(style="thin", color="CCCCCC")] * 4)


def generate(result: ChecklistResult, filename: Optional[str] = None) -> str:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb = Workbook()

    # --- Sheet 1: 제품정보 + 요청사항 ---
    ws = wb.active
    ws.title = "요약"
    _write_title(ws, 1, "우리농 신규물품 자가진단 체크리스트")
    ws["A3"] = "제품 기본정보"
    ws["A3"].font = BOLD
    row = 4
    for k, v in result.product_info.items():
        ws.cell(row, 1, k).font = BOLD
        ws.cell(row, 2, str(v) if v is not None else "")
        row += 1

    row += 1
    ws.cell(row, 1, "생산자 요청사항").font = BOLD
    row += 1
    if result.request_to_producer:
        for i, req in enumerate(result.request_to_producer, 1):
            ws.cell(row, 1, i)
            ws.cell(row, 2, req).alignment = WRAP
            row += 1
    else:
        ws.cell(row, 2, "추가 요청사항 없음")

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 70

    # --- Sheet 2: 필요 서류 ---
    _write_checklist_sheet(wb, "필요서류", result.required_documents)
    # --- Sheet 3: 기재 점검 ---
    _write_checklist_sheet(wb, "기재점검", result.spec_check)
    # --- Sheet 4: 원재료 주의사항 ---
    _write_checklist_sheet(wb, "원재료주의", result.ingredient_warnings)
    # --- Sheet 5: 이슈 대비 ---
    _write_checklist_sheet(wb, "이슈대비", result.potential_issues)

    # 저장
    if not filename:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        pname = result.product_info.get("제품명", "결과").replace(" ", "_") or "결과"
        filename = f"체크리스트_{pname}_{ts}.xlsx"
    path = os.path.join(OUTPUT_DIR, filename)
    wb.save(path)
    return path


def _write_title(ws, row, text):
    c = ws.cell(row, 1, text)
    c.font = WHITE
    c.fill = HEADER_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)


def _write_checklist_sheet(wb, title, items):
    ws = wb.create_sheet(title)
    ws.append(["상태", "아이콘", "제목", "설명", "액션", "근거"])
    for cell in ws[1]:
        cell.font = WHITE
        cell.fill = HEADER_FILL
    for it in items:
        ws.append([it.status, it.icon, it.title, it.description, it.action, it.regulation_ref])
    widths = [12, 6, 35, 60, 40, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = WRAP
